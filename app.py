import re
import os
import streamlit as st
from transformers import pipeline
from openpyxl import Workbook, load_workbook

# Load the classifier model
@st.cache_resource
def load_model():
    return pipeline("text-classification", model="dianapps-vaibhav/distilbert-hiring-intent")

classifier = load_model()

# Validation logic
def validate_answer(answer, q_type):
    if q_type == 'age':
        return answer.isdigit() and 18 <= int(answer) <= 100
    elif q_type == 'gender':
        return answer.lower() in ['male', 'female', 'other']
    elif q_type == 'college':
        return len(answer.strip()) > 2 and all(x.isalpha() or x.isspace() for x in answer.strip())
    elif q_type == 'cgpa':
        return re.match(r'^[0-9](\.[0-9]{1,2})?$', answer) and 0 <= float(answer) <= 10
    elif q_type == 'experience':
        return answer.isdigit()
    elif q_type == 'year':
        return answer.isdigit() and 1 <= int(answer) <= 4
    elif q_type == 'name':
        return len(answer.strip()) > 2 and all(x.isalpha() or x.isspace() for x in answer.strip())
    elif q_type == 'skills':
        return re.match(r'^[a-zA-Z, ]+$', answer)
    elif q_type == 'reason':
        return len(answer.strip()) > 5
    else:
        result = classifier(answer)
        return result[0]['label'] == 'valid'

# Save responses
def save_to_excel(answers, filename="hiring_data.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        row_data = [answers.get(header, "") for header in headers]
        sheet.append(row_data)
    else:
        wb = Workbook()
        sheet = wb.active
        headers = list(answers.keys())
        values = list(answers.values())
        sheet.append(headers)
        sheet.append(values)
    wb.save(filename)

# Question list
questions = [
    ("What is your full name?", "name"),
    ("How old are you?", "age"),
    ("What is your gender? (Male/Female/Other)", "gender"),
    ("What is your college name?", "college"),
    ("What year are you in college? (1-4)", "year"),
    ("What is your CGPA?", "cgpa"),
    ("Years of work experience?", "experience"),
    ("Skills you have (e.g., Python, ML, Java)", "skills"),
    ("Why are you applying for this job?", "reason")
]

# Title
st.title("💬 Hiring Chatbot")
st.markdown("Talk to the hiring assistant below:")

# Initialize session state
if "answers" not in st.session_state:
    st.session_state.answers = {}
if "q_index" not in st.session_state:
    st.session_state.q_index = 0
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []

# Get current question
if st.session_state.q_index < len(questions):
    current_q, current_type = questions[st.session_state.q_index]

# Display previous chat
for entry in st.session_state.chat_history:
    with st.chat_message("user"):
        st.markdown(entry["user"])
    with st.chat_message("ai"):
        st.markdown(entry["bot"])

# Input box
if st.session_state.q_index < len(questions):
    st.markdown(f"**{current_q}**")  # Displays the question on screen
    user_input = st.chat_input("Type your response below...")
    if user_input:
        # Validate
        if validate_answer(user_input, current_type):
            st.session_state.answers[current_type] = user_input
            st.session_state.chat_history.append({
                "user": user_input,
                "bot": "✅ Noted! Let's move on."
            })
            st.session_state.q_index += 1
        else:
            st.session_state.chat_history.append({
                "user": user_input,
                "bot": "❌ Invalid response. Please try again."
            })
        st.rerun()
else:
        # Save data
        save_to_excel(st.session_state.answers)

        # Show confirmation
        with st.chat_message("ai"):
            st.success("✅ Your responses have been saved successfully!")
            st.markdown("### 📋 Your Responses:")

            # Display answers nicely
            for key, value in st.session_state.answers.items():
                label = key.capitalize().replace("_", " ")
                st.markdown(f"- **{label}**: {value}")

            # Retake option
            if st.button("🔄 Retake the Form"):
                st.session_state.answers = {}
                st.session_state.q_index = 0
                st.session_state.chat_history = []
                st.rerun()
